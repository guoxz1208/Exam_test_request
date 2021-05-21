#!/usr/bin/env python
# -*- coding:utf-8 -*-
# fileName: handle_excel.py
"""
通过xlrd读Excel文件就行读取，openpyxl对Excel文件进行写入
"""
import xlrd,xlwt
import openpyxl
import os,sys
base_path = os.path.abspath(os.path.join(os.getcwd(),'../'))
sys.path.append(base_path)
from common.get_log import get_log
# logger = get_log()

class HandleExcel:
    def __init__(self):
        self.logger = get_log()

    def read_excel(self,fileName=None):
        # self.logger.info("==========读取Excel==========")
        if fileName == None:
            Excel_Path = base_path + '/data/接口测试用例-v1.0.xlsx'
        else:
            Excel_Path = base_path + fileName
        Excel_Data = xlrd.open_workbook(Excel_Path)
        return Excel_Data

    def excel_sheet(self,SheetNum=None,fileName=None):
        # self.logger.info("==========读取Excel的sheet页==========")
        """
        :param SheetNum: sheet页数，如第几个sheet
        :param fileName: 文件路径
        """
        sheet_add = self.read_excel(fileName)
        if SheetNum == None:
            SheetNum = 1
        Sheet_Name = sheet_add.sheet_by_index(SheetNum)
        return Sheet_Name

    def row_excel(self,SheetNum=None,fileName=None):        # 获取总行数
        # self.logger.info("==========读取Excel总行数==========")
        """
        :param SheetNum: sheet页数，如第几个sheet
        """
        row_value = self.excel_sheet(SheetNum,fileName).nrows
        return row_value

    def row_data_excel(self,rowNum,SheetNum=None,fileName=None):        # 获取一行全部内容
        # self.logger.info("==========读取Excel第："+str(rowNum)+"行内容==========")
        """
        :param RowNum: 行数
        """
        row_data = self.excel_sheet(SheetNum,fileName).row_values(rowNum)
        return row_data

    def cols_excel(self,SheetNum=None,fileName=None):        # 获取总列数
        # self.logger.info("==========读取Excel总列数==========")
        cols_value = self.excel_sheet(SheetNum,fileName).ncols
        return cols_value

    def cols_data_excel(self,colNum,SheetNum=None,fileName=None):        # 获取一列全部内容
        # self.logger.info("==========读取Excel第："+str(colNum)+"列内容==========")
        """
        :param colNum: 列数
        """
        col_data = self.excel_sheet(SheetNum,fileName).col_values(colNum)
        return col_data

    def cell_data_excel(self,rowNum,colNum,SheetNum=None,fileName=None):        # 获取某一个单元格内容
        # self.logger.info("==========读取Excel某一个单元格内容==========")
        cell_data = self.excel_sheet(SheetNum,fileName).cell_value(rowNum,colNum)
        return cell_data

    def all_data_excel(self):        # 获取Excel全部内容
        # self.logger.info("==========读取Excel全部文件内容操作==========")
        rowList = []
        row_num = self.row_excel()

        for i in range(row_num):
            all_data = self.row_data_excel(i)
            rowList.append(all_data)
        return rowList

    def write_excle(self,index, row, column, value=None,fileName=None):        # 使用openpyxl对Excel写入内容
        # self.logger.info("==========Excel写入文件操作==========")
        if fileName == None:
            Excel_Path = base_path + '/data/接口测试用例-v1.0.xlsx'
        else:
            Excel_Path = base_path + fileName
        wb = openpyxl.load_workbook(Excel_Path)
        sheet = wb.sheetnames
        ws = wb[sheet[index]]
        try:
            if value is not None:
                ws.cell(row,column,value)
                self.logger.info("==========Excel文件内容写入成功==========")
            else:
                self.logger.debug("==========Excel文件内容写入失败==========")
        except:
            self.logger.debug("========文件内容写入失败========")
        wb.save(Excel_Path)

if __name__ == '__main__':
    excel_add = HandleExcel()
    # print(excel_add.cols_excel(1))
    # print(excel_add.cols_data_excel(0))
    # print(excel_add.cell_data_excel(2,2))
    res = excel_add.all_data_excel()
    for one in res:
        print(one)
    # excel_add.write_excle(0,10,1,'test')